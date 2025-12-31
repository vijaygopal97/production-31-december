#!/usr/bin/env python3
import pandas as pd
import json
import sys

def is_yellow_fill(cell):
    """Check if a cell has yellow background fill"""
    if not cell.fill or not cell.fill.start_color:
        return False
    
    # Check for yellow fill - yellow is typically RGB(255, 255, 0) or similar
    # openpyxl stores colors in different formats, check for yellow patterns
    fill_color = cell.fill.start_color
    if hasattr(fill_color, 'rgb') and fill_color.rgb:
        # Check if RGB contains yellow (high red and green, low blue)
        rgb = fill_color.rgb
        if rgb and len(rgb) >= 6:  # RGB hex string like 'FFFFFF00'
            # Extract RGB values (skip 'FF' alpha if present)
            rgb_hex = rgb[-6:] if len(rgb) > 6 else rgb
            r = int(rgb_hex[0:2], 16)
            g = int(rgb_hex[2:4], 16)
            b = int(rgb_hex[4:6], 16)
            # Yellow: high red and green, low blue (with some tolerance)
            if r > 200 and g > 200 and b < 100:
                return True
    
    # Also check for theme colors or indexed colors that might be yellow
    if hasattr(fill_color, 'theme') or hasattr(fill_color, 'index'):
        # Check if it's a known yellow index (varies by Excel version)
        # Common yellow index is around 43-44
        if hasattr(fill_color, 'index') and fill_color.index in [43, 44, 45]:
            return True
    
    return False

def generate_ac_master_data():
    """Generate JSON file with AC, MP, and MLA data from Excel"""
    excel_file = '/var/www/West_Bengal_State_Master_-_2025 v3 (1).xlsx'
    output_file = '/var/www/opine/backend/data/acMasterData.json'
    
    try:
        # Read Excel file using openpyxl to get exact cell values and formatting
        import openpyxl
        # Use data_only=False to preserve formatting information
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        sheet = wb.active
        
        ac_data_map = {}
        bye_election_ac_names = set()  # Store AC names that have bye-elections
        
        # Find AC No column by searching header row (row 3)
        ac_no_column = None
        header_row = 3
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and 'AC No' in str(cell_value):
                ac_no_column = col
                break
        
        if not ac_no_column:
            print("⚠️ Warning: Could not find 'AC No' column. Bye-election detection will be skipped.")
        
        # Row 3 = Header (index 3 in 1-based)
        # Row 5 onwards = Data (index 5 in 1-based)
        # Column C = AC Name (column 3)
        # Column H = Current MLA - 2021 (column 8)
        # Column DY = 2024 MP (column 129)
        
        row = 5
        while True:
            ac_cell = sheet.cell(row=row, column=3)  # Column C - AC Name
            if not ac_cell.value:
                break
            
            ac_name = str(ac_cell.value).strip()
            # Skip if empty or numeric (summary rows)
            if not ac_name or ac_name == 'None' or ac_name == 'null' or ac_name.isdigit():
                row += 1
                continue
            
            # Get MLA name from column H (column 8)
            mla_cell = sheet.cell(row=row, column=8)
            mla_name = None
            if mla_cell.value:
                mla_name_str = str(mla_cell.value).strip()
                if mla_name_str and mla_name_str != 'None' and mla_name_str != 'null' and not mla_name_str.replace('.', '').isdigit():
                    mla_name = mla_name_str
            
            # Get MP name from column DY (column 129)
            mp_cell = sheet.cell(row=row, column=129)
            mp_name = None
            if mp_cell.value:
                mp_name_str = str(mp_cell.value).strip()
                if mp_name_str and mp_name_str != 'None' and mp_name_str != 'null' and not mp_name_str.replace('.', '').isdigit():
                    mp_name = mp_name_str
            
            # Check if AC No cell is highlighted in yellow (bye-election AC)
            has_bye_election = False
            if ac_no_column:
                ac_no_cell = sheet.cell(row=row, column=ac_no_column)
                if is_yellow_fill(ac_no_cell):
                    has_bye_election = True
                    bye_election_ac_names.add(ac_name)
            
            if ac_name:
                ac_data_map[ac_name] = {
                    'mpName': mp_name,
                    'mlaName': mla_name,
                    'hasByeElection': has_bye_election
                }
            
            row += 1
        
        # Save to JSON file
        import os
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(ac_data_map, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Generated AC master data: {len(ac_data_map)} ACs")
        print(f"📁 Saved to: {output_file}")
        print(f"📋 Sample ACs: {list(ac_data_map.keys())[:5]}")
        print(f"📋 Bye-election ACs: {len(bye_election_ac_names)} ACs")
        if bye_election_ac_names:
            print(f"📋 Bye-election AC names: {list(bye_election_ac_names)[:10]}")
        
        return ac_data_map
        
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    generate_ac_master_data()




